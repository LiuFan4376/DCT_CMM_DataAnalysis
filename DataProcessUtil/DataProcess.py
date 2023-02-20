import pandas as pd
import openpyxl
import os


class DataAnalysis:
    data_path = ''
    module_path = ''
    data_clean_path = ''
    output_path = ''

    def __init__(self, data__path, module_path, data_clean_path, output_path):
        self.data_path = data__path
        self.module_path = module_path
        self.data_clean_path = data_clean_path
        self.output_path = output_path

    # 数据清洗方法
    def data_clean(self):
        data_number = 420  # 要插入的记录数
        # 读取excel文件
        data_sheet = pd.read_excel(self.data_path, sheet_name=0, skiprows=range(0, 10))
        colums_list = data_sheet.columns.values
        # 删除多余的列
        del_colums = [colums_list[3]] + [colums_list[5]] + [colums_list[10]] + list(colums_list[12:len(colums_list)])
        data_sheet = data_sheet.drop(columns=del_colums)
        data_sheet.dropna(axis=1, how='all', inplace=True)
        # 删除多余的行
        data_sheet.dropna(axis=0, how='any', inplace=True)
        for row_index in data_sheet.index.values:
            temp1 = data_sheet.loc[row_index, '轴']
            temp2 = data_sheet.loc[row_index, colums_list[1]]
            # print(temp2[0:2])
            if temp1 == '轴':
                data_sheet.drop(index=row_index, inplace=True)
            if row_index > 18 and temp2[0:2] == '位置':
                data_sheet.drop(index=row_index, inplace=True)
        # 交换部分索引位置

        data_sheet.drop(index=data_sheet.index[data_number:data_sheet.shape[0]], inplace=True)
        data_sheet.to_excel(self.data_clean_path)
        return

    # 数据拷贝到模板文件方法
    def data_copy(self):
        # 获取模板工作薄
        wb_module_path = self.module_path
        wb_module = openpyxl.load_workbook(wb_module_path)
        sheet_module_A = wb_module['A測定']
        wb_data = openpyxl.load_workbook(self.data_clean_path)
        # print(wb_data.sheetnames)
        data_sheet = wb_data['Sheet1']
        data_number = 420  # 要插入的记录数
        print('模版化文件正在输出....')
        for i in range(1, data_number + 1):
            # print(data_sheet.cell(i, 7).value)
            if i == 4 or i == 5:
                print(data_sheet.cell(10 - i, 6).value)
                sheet_module_A.cell(11 + i, 11).value = float(data_sheet.cell(i + 3, 6).value)
            elif i == 6 or i == 7:
                sheet_module_A.cell(11 + i, 11).value = float(data_sheet.cell(i - 1, 6).value)
                sheet_module_A.cell(11 + i, 11).value = float(data_sheet.cell(i - 1, 6).value)
            else:
                sheet_module_A.cell(11 + i, 11).value = float(data_sheet.cell(i + 1, 6).value)
        module_sheet_name = wb_module.sheetnames
        for names in module_sheet_name:
            if names != 'A測定':
                del wb_module[names]
        os.chdir(self.output_path)
        wb_module.save('A程序模板化.xlsx')
        print('输出完成')


def data_processB():
    import openpyxl

    '''
    获取三坐标检测数据EXCEL文件
    '''
    wb_path = input("请输入三坐标检测数据路径：")
    # wb_path = '/Users/liufan/Desktop/DC调试计划/GB5R051220608A0050-B-上机确认-001.XLSX'
    wb = openpyxl.load_workbook(wb_path)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]  # 获取指定sheet表

    '''
    处理三坐标检测数据表
    '''
    print('正在初始化数据文件...')
    sheet.delete_rows(1, 9)  # 删除表头信息
    sheet.delete_cols(1, 1)
    max_row = sheet.max_row
    max_column = sheet.max_column

    # 删除多余行
    # 删除多余行
    i = 1
    while i <= max_row:
        temp = sheet.cell(i, 2).value
        loc = sheet.cell(i, 1).value

        if temp is None or temp == '单位' or temp == '描述':
            sheet.delete_rows(i)
            max_row = max_row - 1
        elif i > 7 and loc[0:2] == '位置':
            print(loc)
            if loc != '':
                sheet.delete_rows(i)
                max_row = max_row - 1
        else:
            i = i + 1

    '''
    获取模板工作薄
    '''
    wb_module_path = input("请输入数据整理模版路径：")
    # wb_module_path = '/Users/liufan/Desktop/DC调试计划/模板.xlsx'
    wb_module = openpyxl.load_workbook(wb_module_path)
    sheet_module_B = wb_module['B測定']
    data_number = 142  # 要插入的记录数
    print('模版化文件正在输出....')
    '''
    将检测数据表中的记录统计到模板文件中
    '''
    j = 0
    for i in range(1, data_number):
        j = j + 1
        print(sheet.cell(j, 7).value)
        sheet_module_B.cell(11 + i, 11).value = float(sheet.cell(j, 7).value)

    # output_path=input('请输入处理后的工作表存储路径：')
    output_path = '/Users/liufan/Desktop'
    # os.chdir(output_path)  # 修改工作路径
    output = output_path + '/B程序模板化.xlsx'
    wb_module.save(output)
    print('输出完成')
    wb.save('/Users/liufan/Desktop/B_clean.xlsx')
